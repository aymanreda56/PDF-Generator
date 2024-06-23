import customtkinter as ctk
import re, math
from datetime import date, timedelta

import openpyxl.styles
from enums import EntryError, EntryErrorCode, ArmyLevels
# from MainMenu import MainMenu
import helpers
import os
import pandas as pd
import numpy as np
from style import *
import openpyxl

FONT_STYLE = 'Dubai'




# class VacationsHistory():



#     def Draw_Big_Scrollable_Frame(self):

#         all_present_soldiers = helpers.fetchSoldiers()
#         data_to_be_displayed = []
#         for soldier in all_present_soldiers:
#             Soldier_ID = soldier['Soldier_ID']
#             Soldier_Name = soldier['Name']
#             last_vac_date = helpers.getLastReturnFromID(Soldier_ID=Soldier_ID)
#             service_days = helpers.getServiceDays(last_vac_date)
            

#             data_to_be_displayed.append((Soldier_Name, last_vac_date, service_days, Soldier_ID))
        
#         # all_absent_soldiers = helpers.getActiveVacations(with_disabled=True)
#         # for soldier in all_absent_soldiers:
#         #     Soldier_ID = soldier[0]
#         #     Soldier_Name = soldier[1]
#         #     last_vac_date = helpers.getLastReturnFromID(Soldier_ID=Soldier_ID)
#         #     service_days = helpers.getServiceDays(last_vac_date)
#         #     data_to_be_displayed.append((Soldier_Name, last_vac_date, service_days, Soldier_ID))


#         data_to_be_displayed.sort(key=lambda x: x[2], reverse= True) #sort according to service days


#         frame_for_scrollable_frame = ctk.CTkFrame(self.root, width=1460,fg_color=FRAME_DARK_COLOR, border_color=BUTTON_COLOR, border_width=5, corner_radius=15)
#         frame_for_scrollable_frame.pack()
        
#         Header_Frame = ctk.CTkFrame(master=frame_for_scrollable_frame, fg_color=FRAME_LIGHT_COLOR, width=1460, height=30, corner_radius=5)
#         Header_Frame.pack(pady=8, padx=8)
        
        
#         Master_Frame = ctk.CTkScrollableFrame(frame_for_scrollable_frame, label_text="", width=1430, height=500,fg_color=FRAME_DARK_COLOR, label_fg_color=FRAME_DARK_COLOR, scrollbar_button_color=BUTTON_COLOR)
#         # self.big_Entire_Frame = entire_preview_frame

#         # Master_Frame = ctk.CTkScrollableFrame(master=self.root, width=1000, height=500)
#         Master_Frame.pack(padx=10)

        

#         Service_Days_Lbl = ctk.CTkLabel(master=Header_Frame, text= 'عدد أيام العمل', font=(FONT_STYLE, 16, 'bold'), justify = ctk.RIGHT, text_color=FG_COLOR)
#         Service_Days_Lbl.place(relx=0.3, rely=0.5, anchor='center')

#         Last_Date_Lbl = ctk.CTkLabel(master=Header_Frame, text= 'تاريخ آخر عودة', font=(FONT_STYLE, 16, 'bold'), justify = ctk.RIGHT, text_color=FG_COLOR)
#         Last_Date_Lbl.place(relx=0.6, rely=0.5, anchor='center')


#         Name_Lbl = ctk.CTkLabel(master=Header_Frame, text= 'الإسم', font=(FONT_STYLE, 16, 'bold'), justify = ctk.RIGHT, text_color=FG_COLOR)
#         Name_Lbl.place(relx=0.88, rely=0.5, anchor='center')


#         for soldier in data_to_be_displayed:
#             self.Add_Row(Master_Frame, soldier)




#     def __init__(self):
#         self.root = ctk.CTk(fg_color=BG_COLOR)

#         self.root.title("تنظيم و أفراد مكتب السيد/ مدير الجهاز")
#         screen_width, screen_height = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
#         width, height = 1700, 900
#         self.root.geometry(f"{width}x{height}+{str(math.floor(screen_width/2 - width/2))}+{str(math.floor(screen_height/2 - height/2))}")  # Set window size
#         self.root.iconbitmap("../data/icolog.ico")

#         BigLabel = ctk.CTkLabel(master=self.root, text='عرض كل الأجازات', font=(FONT_STYLE, 60, 'bold'),text_color=TEXT_COLOR)
#         BigLabel.pack()

#         self.root.mainloop()




# VacationsHistory()


        


def preprocessing():
    all_soldiers = helpers.fetchSoldiers()
    all_vacations_history = helpers.Get_All_Vacations_History()


    names = ['الإسم/التاريخ']
    names.append('اليوم')
    the_very_first_date = date(year=2100, month=1, day=1)

    #filling the list of names and getting the very first date of vacation
    for sold in all_soldiers:
        names.append(sold['Name'])
    for i, vacation in enumerate(all_vacations_history):
        if(date.fromisoformat(vacation['From_Date']) < the_very_first_date):
            the_very_first_date = date.fromisoformat(vacation['From_Date'])

    #getting the number of days from the very first date till today
    d0 = the_very_first_date
    d1 = date.today()
    delta = d1 - d0
    num_of_days = delta.days

    all_history_np = np.ndarray(shape=(num_of_days + 1, len(names)), dtype=str) #the +3 and +1 to compensate for extra columns like date and weekday


    df = pd.DataFrame(all_history_np, index=None)

    return all_soldiers, df, names, the_very_first_date



#prepopulating the headers
def prepopulate_headers(df:pd.DataFrame, names_list:list, the_very_first_date:date):
    
    days_map = {0: 'الإثنين', 1: "الثلاثاء", 2: "الأربعاء", 3: "الخميس", 4: "الجمعة", 5: "السبت", 6:"الأحد"}

    i = 0
    for col in range(0, df.shape[1]):
        df.iloc[0, col] = names_list[i] #the +2 and -2 to compensate for the first two columns corresponding to date and weekday
        i+=1

    for row in range(df.shape[0] - 1):
        date_of_row = the_very_first_date + timedelta(days=row)
        df.iloc[row + 1, 0] = date_of_row.isoformat()
        df.iloc[row + 1, 1] = days_map[date_of_row.weekday()]
    
    #fix the headers
    new_header = df.iloc[0] #grab the first row for the header
    df = df[1:] #take the data less the header row
    df.columns = new_header #set the header row as the df header

    return df



# df.set_index(df.columns[0], inplace=True)
# df.drop(df.columns[0], axis=1, inplace=True)

def populate_sheet(df:pd.DataFrame, names_list:list, all_soldiers_data:list):
    for row in range(0, df.shape[0]):
        if(pd.isna(df.iloc[row, 0])):
            continue
        
        
        for i, name in enumerate(names_list):
            if(i<2): continue
            # print(df.iloc[row, 0])
            vac_code = helpers.TestIfVacation(Soldier_ID=helpers.getSoldierIDFromName(name, all_soldiers_data), date_of_test=df.iloc[row, 0])
            
            # entry = 'موجود'
            if vac_code == 0:
                entry = 'موجود'
            elif vac_code == 1:
                entry = 'أجازة'
            elif vac_code == 2:
                entry = 'إمتداد'
            
            df.iloc[row, list(df.columns).index(name)] = entry
            
    return df








def change_colors_in_excel_sheet(sheet_path:str, num_rows:int, num_cols:int)->None:
    workBook = openpyxl.load_workbook(sheet_path)


    worksheet = workBook.active

    redFill = openpyxl.styles.PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
    
    greenFill = openpyxl.styles.PatternFill(start_color='5feb28',
                   end_color='5feb28',
                   fill_type='solid')
    
    yellowFill = openpyxl.styles.PatternFill(start_color='ffdc2b',
                   end_color='ffdc2b',
                   fill_type='solid')
    
    thin_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'),
            right=openpyxl.styles.borders.Side(style='thin'))
    
    bigfontStyle = openpyxl.styles.Font(name = 'custom', size = "16", bold=True)

    for row in range(1, num_rows+2):
        for col in range(1, num_cols+2):
            
            cell = worksheet.cell(row=row,column=col)
            if(row == 1 or col == 2):
                cell.font = bigfontStyle


            if(cell.value == 'موجود'):
                cell.fill = greenFill
            elif(cell.value == 'أجازة'):
                cell.fill = redFill
            elif(cell.value == 'إمتداد'):
                cell.fill = yellowFill
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            cell.border = thin_border

    column_widths = []
    
    for col in range(1, num_cols+2):
        max_col_size = 0
        for row in range(1, num_rows+2):
            cell = worksheet.cell(row=row,column=col)
            try:
                max_col_size = max(len(cell.value) * 1.3, max_col_size)
            except: max_col_size = 10
        
        column_widths.append(max_col_size)
        
    for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width


    worksheet.sheet_view.rightToLeft = True
    
    workBook.save(sheet_path)  # save the workbook
    workBook.close()  # close the workbook

   


def ExportVacationsHistory():
    all_soldiers, df, names, the_very_first_date = preprocessing()
    df = prepopulate_headers(df=df, names_list=names, the_very_first_date=the_very_first_date)
    df = populate_sheet(df=df, names_list=names, all_soldiers_data=all_soldiers)

    df.to_excel('Vacations_History.xlsx')


    change_colors_in_excel_sheet('Vacations_History.xlsx', num_rows=df.shape[0], num_cols=df.shape[1])


    os.startfile('Vacations_History.xlsx')






